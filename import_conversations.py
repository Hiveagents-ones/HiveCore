#!/usr/bin/env python
"""
导入 Excel 对话数据到后端数据库

四个对话硬编码 ID：
- ID 1: 上传文档
- ID 2: 读取记忆
- ID 3: 社区论坛
- ID 4: 无数据来源
"""

import os
import sys
import django

# 设置 Django 环境
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "backend.settings")
django.setup()

from api.models import Conversation, Message
import openpyxl


# Excel 工作表名称到对话 ID 的映射
SHEET_TO_ID = {
    "上传文档": 1,
    "记忆": 2,
    "社区": 3,
    "无数据来源": 4,
}

# 对话标题
ID_TO_TITLE = {
    1: "上传文档",
    2: "读取记忆",
    3: "社区论坛",
    4: "无数据来源",
}

EXCEL_PATH = "/Users/prayer/yiju/mock-data/假数据-1.xlsx"


def parse_messages_from_sheet(ws):
    """从工作表中解析对话消息"""
    messages = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        user_msg = row[0]  # 第一列：用户消息
        assistant_msg = row[1]  # 第二列：Assistant 回复

        # 遇到分隔符停止
        if user_msg and "=====" in str(user_msg):
            break

        # 添加用户消息
        if user_msg and str(user_msg).strip():
            messages.append({
                "type": "user",
                "content": str(user_msg).strip(),
            })

        # 添加 Assistant 回复
        if assistant_msg and str(assistant_msg).strip():
            messages.append({
                "type": "assistant",
                "content": str(assistant_msg).strip(),
            })

    return messages


def import_conversations():
    """导入所有对话"""
    # 加载 Excel 文件
    print(f"加载 Excel 文件: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    for sheet_name, conv_id in SHEET_TO_ID.items():
        print(f"\n处理工作表: {sheet_name} -> 对话 ID: {conv_id}")

        if sheet_name not in wb.sheetnames:
            print(f"  警告: 工作表 '{sheet_name}' 不存在，跳过")
            continue

        ws = wb[sheet_name]
        messages = parse_messages_from_sheet(ws)

        if not messages:
            print(f"  警告: 没有找到消息，跳过")
            continue

        # 删除现有的对话（如果存在）
        Conversation.objects.filter(id=conv_id).delete()

        # 创建对话，指定 ID
        conv = Conversation(
            id=conv_id,
            title=ID_TO_TITLE[conv_id],
        )
        conv.save()

        # 重置自增序列（PostgreSQL）或直接插入（SQLite）
        # Django 会自动处理

        print(f"  创建对话: ID={conv.id}, 标题={conv.title}")

        # 创建消息
        for idx, msg_data in enumerate(messages):
            msg = Message.objects.create(
                conversation=conv,
                type=msg_data["type"],
                content=msg_data["content"],
            )
            print(f"    消息 {idx + 1}: [{msg_data['type']}] {msg_data['content'][:50]}...")

        print(f"  导入 {len(messages)} 条消息")

    print("\n导入完成!")

    # 显示结果
    print("\n当前对话列表:")
    for conv in Conversation.objects.all().order_by("id"):
        msg_count = conv.messages.count()
        first_msg = conv.messages.filter(type="user").first()
        first_content = first_msg.content[:50] if first_msg else "无"
        print(f"  ID {conv.id}: {conv.title} ({msg_count} 条消息)")
        print(f"    首条用户消息: {first_content}...")


if __name__ == "__main__":
    import_conversations()
